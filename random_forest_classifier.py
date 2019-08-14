import matplotlib 
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import numpy as np
from numpy.linalg import pinv
from pandas import DataFrame
import pandas as pd
import random
from sklearn import model_selection#from sklearn import mixture
from sklearn import svm
from sklearn.ensemble import RandomForestClassifier
from sklearn.datasets import make_classification 
import sys
import numpy

def visualize_classifier(model, X, y, ax=None, cmap='rainbow'):
    ax = ax or plt.gca()
    
    # Plot the training points
    ax.scatter(X[:, 0], X[:, 1], c=y, s=30, cmap=cmap,
               clim=(y.min(), y.max()), zorder=3)
    ax.axis('tight')
    ax.axis('off')
    xlim = ax.get_xlim()
    ylim = ax.get_ylim()
    
    # fit the estimator
    model.fit(X, y)
    xx, yy = np.meshgrid(np.linspace(*xlim, num=200),
                         np.linspace(*ylim, num=200))
    Z = model.predict(np.c_[xx.ravel(), yy.ravel()]).reshape(xx.shape)

    # Create a color plot with the results
    n_classes = len(np.unique(y))
    contours = ax.contourf(xx, yy, Z, alpha=0.3,
                           levels=np.arange(n_classes + 1) - 0.5,
                           cmap=cmap, clim=(y.min(), y.max()),
                           zorder=1)

    ax.set(xlim=xlim, ylim=ylim)

def readExcelSheet1(excelfile):
    from pandas import read_excel
    return (read_excel(excelfile)).values

def readExcelRange(excelfile,sheetname="Sheet1",startrow=1,endrow=1,startcol=1,endcol=1):
    from pandas import read_excel
    values=(read_excel(excelfile, sheetname,header=None)).values;
    return values[startrow-1:endrow,startcol-1:endcol]

def readExcel(excelfile,**args):
    if args:
        data=readExcelRange(excelfile,**args)
    else:
        data=readExcelSheet1(excelfile)
    if data.shape==(1,1):
        return data[0,0]
    elif (data.shape)[0]==1:
        return data[0]
    else:
        return data

def writeExcelData(x,excelfile,sheetname,startrow,startcol):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df=DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname,startrow=startrow-1, startcol=startcol-1, header=False, index=False)
    writer.save()
    writer.close()


def writeExcelDataHeader(x,excelfile,sheetname,startrow,startcol):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df=DataFrame(x)
    book = load_workbook(excelfile)
    writer = ExcelWriter(excelfile, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheetname,startrow=startrow-1, startcol=startcol-1, header=True, index=False)
    writer.save()
    writer.close()


def getSheetNames(excelfile):
    from pandas import ExcelFile
    return (ExcelFile(excelfile)).sheet_names







excelfile = '/pathtothisfile/final_master_ryan.xlsx'
data = readExcel(excelfile)
print 'full data shape: ', data.shape
D_train, D_test = model_selection.train_test_split(data, test_size=0.3)
print 'D_train Shape: ', D_train.shape
print 'D_test Shape: ', D_test.shape



D_test_class_list = D_test[:, -1]


D_training_class_list = D_train[:, -1]

D_training_without_classes = D_train[:, :-1]


D_test_features = D_test[:, :-1]



cleaned_D_test_features = np.nan_to_num(D_test_features)

cleaned_D_training = np.nan_to_num(D_training_without_classes)

clf = RandomForestClassifier(n_estimators=100, bootstrap=False)
#clf = RandomForestClassifier(n_estimators=100, min_samples_leaf=10, bootstrap=False)
clf.fit(cleaned_D_training, D_training_class_list)
print 'IMPORTANCES:\n ', clf.feature_importances_

#visualize_classifier(clf,cleaned_D_training, D_training_class_list)

print 'Sum of feature importances: ', np.sum(clf.feature_importances_)

predictions = clf.predict(cleaned_D_test_features)

#numpy.set_printoptions(threshold=sys.maxsize)


def randomCross():
	#THIS IS USING CROSS VALIDATION
	from sklearn.model_selection import RandomizedSearchCV
	# Number of trees in random forest
	n_estimators = [int(x) for x in np.linspace(start = 200, stop = 2000, num = 10)]
	# Number of features to consider at every split
	max_features = ['auto', 'sqrt']
	# Maximum number of levels in tree
	max_depth = [int(x) for x in np.linspace(10, 110, num = 11)]
	max_depth.append(None)
	# Minimum number of samples required to split a node
	min_samples_split = [2, 5, 10]
	# Minimum number of samples required at each leaf node
	min_samples_leaf = [1, 2, 4]
	# Method of selecting samples for training each tree
	bootstrap = [True, False]
	# Create the random grid
	random_grid = {'n_estimators': n_estimators,
               'max_features': max_features,
               'max_depth': max_depth,
               'min_samples_split': min_samples_split,
               'min_samples_leaf': min_samples_leaf,
               'bootstrap': bootstrap}


	# Use the random grid to search for best hyperparameters
	# First create the base model to tune
	rf = RandomForestClassifier()
	# Random search of parameters, using 3 fold cross validation, 
	# search across 100 different combinations, and use all available cores
	rf_random = RandomizedSearchCV(estimator = rf, param_distributions = random_grid, n_iter = 100, cv = 3, verbose=2, random_state=42, n_jobs = -1)
	# Fit the random search model
	rf_random.fit(cleaned_D_training, D_training_class_list)

	print 'These are rf best params: '
	print rf_random.best_params_


	best_random = rf_random.best_estimator_
	random_predictions = best_random.predict(cleaned_D_test_features)
	#random_predictions =  evaluate(best_random, cleaned_D_test_features, D_test_class_list)

	random_predictions_list = random_predictions.tolist()
	correct_random_predictions = 0

	for i in range(len(actuals_list)):
		if float(random_predictions_list[i] == float(actuals_list[i])):
			correct_random_predictions = correct_random_predictions + 1

	
	print 'Correct random predictions: ', correct_random_predictions


	random_accuracy = float(correct_random_predictions)/2514
	print 'This is the accuracy of random: ', random_accuracy



def gridCross(cleaned_D_training, D_training_class_list,cleaned_D_test_features, actuals_list ):
	from sklearn.model_selection import GridSearchCV
	# Create the parameter grid based on the results of random search 
	param_grid = {
    	'bootstrap': [True, False],
    	'max_depth': [60, 70, 80, 90],
    	'max_features': ['sqrt', 'auto'],
    	'min_samples_leaf': [1, 2, 3],
    	'min_samples_split': [2,3,4],
    	'n_estimators': [65, 100, 500]
		}
	# Create a based model
	rf = RandomForestClassifier()
	# Instantiate the grid search model
	grid_search = GridSearchCV(estimator = rf, param_grid = param_grid, 
                          cv = 3, n_jobs = -1, verbose = 2)

	grid_search.fit(cleaned_D_training, D_training_class_list)

	print 'These are grid_search best params: '
	print grid_search.best_params_


	best_grid = grid_search.best_estimator_
	grid_predictions = best_grid.predict(cleaned_D_test_features)
	#random_predictions =  evaluate(best_random, cleaned_D_test_features, D_test_class_list)

	grid_predictions_list = grid_predictions.tolist()
	correct_grid_predictions = 0

	for i in range(len(actuals_list)):
		if float(grid_predictions_list[i] == float(actuals_list[i])):
			correct_grid_predictions = correct_grid_predictions + 1

	
	print 'Correct crossgrid predictions: ', correct_grid_predictions


	grid_accuracy = float(correct_grid_predictions)/2514
	print 'This is the accuracy of grid_search: ', grid_accuracy









predictions_list = predictions.tolist()
actuals_list = D_test_class_list.tolist()



correct_predictions = 0

#randomCross(cleaned_D_training, D_training_class_list,cleaned_D_test_features, actuals_list)
#gridCross(cleaned_D_training, D_training_class_list,cleaned_D_test_features, actuals_list)

for i in range(len(actuals_list)):
	if float(predictions_list[i]) == float(actuals_list[i]):
		correct_predictions = correct_predictions + 1

print 'Correct predictions: ',correct_predictions


accurarcy = float(correct_predictions)/2514



print 'This is the accuracy of base: ', accurarcy

def plotData(x_data, y_data):
	plt.scatter(x_data,y_data)
	plt.plot(x_data,y_data)
	plt.xlabel('Max Depth W/ n_estimators = 100')
	plt.ylabel('Prediction Accuracy')
	plt.show()

xplot_numtrees = [10, 50, 100, 200, 300, 400, 500] #this is just for base runs
yplot_accuracy = [0.849244232299, 0.860779634049, 0.876690533015, 0.869530628481, 0.866348448687, 0.867143993636, 0.86754176611]

plotData(xplot_numtrees, yplot_accuracy)

xplot_maxdepth = [2, 5, 10, 20, 40, 50]
yplot_accuracydepth = [0.826968973747, 0.837311058075 , 0.854017501989, 0.861177406523, 0.867143993636, 0.865155131265]

plotData(xplot_maxdepth, yplot_accuracydepth)


xplot_minsampleleafs = [1, 3, 10, 20, 50]
yplot_minsampleleafs = [0.855210819411,0.865155131265, 0.865155131265, 0.86754176611, 0.858392999204]

plotData(xplot_minsampleleafs,yplot_minsampleleafs )

########################TESTING WITH DROPPING ALL BAD DATA##############



data2 = data[~np.isnan(data).any(axis=1)]
print 'data2 shape: ', data2.shape


D_train2, D_test2 = model_selection.train_test_split(data2, test_size=0.3)
print 'D_train2 Shape: ', D_train2.shape
print 'D_test2 Shape: ', D_test2.shape




D_test_class_list2 = D_test2[:, -1]


D_training_class_list2 = D_train2[:, -1]

D_training_without_classes2 = D_train2[:, :-1]


D_test_features2 = D_test2[:, :-1]

clf2 = RandomForestClassifier(n_estimators=100, min_samples_leaf = 1, bootstrap=False)
#clf = RandomForestClassifier(n_estimators=100, min_samples_leaf = 4, min_samples_split=70, bootstrap=False)
clf2.fit(D_training_without_classes2, D_training_class_list2)
print 'IMPORTANCES:\n ', clf2.feature_importances_


print 'Sum of feature importances: ', np.sum(clf2.feature_importances_)

predictions2 = clf2.predict(D_test_features2)

predictions_list2 = predictions2.tolist()
actuals_list2 = D_test_class_list2.tolist()

correct_predictions2 = 0

#randomCross(cleaned_D_training, D_training_class_list,cleaned_D_test_features, actuals_list)
#gridCross(D_training_without_classes2, D_training_class_list2,D_test_features2, actuals_list2)

for i in range(len(actuals_list2)):
	if float(predictions_list2[i]) == float(actuals_list2[i]):
		correct_predictions2 = correct_predictions2 + 1

print 'Correct predictions2: ',correct_predictions2
print 'Total actuals2 len: ', len(actuals_list2)

accurarcy2 = float(correct_predictions2)/len(actuals_list2)



print 'This is the accuracy2 of base: ', accurarcy2





